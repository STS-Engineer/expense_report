from flask import Flask, request, send_file, render_template_string, jsonify
import os
from PyPDF2 import PdfReader
import pandas as pd
import requests
import fitz  # PyMuPDF for extracting images
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'

# Live currency API
FIXER_API_URL = "http://data.fixer.io/api/latest"
FIXER_API_KEY = "653aca7bac0ce92affcdcb0116ecbc0a"


@app.route('/')
def home():
    return '''
    <!doctype html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>PDF to Excel Converter</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                background-color: #f4f4f9;
                margin: 0;
                padding: 20px;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
            }
            .container {
                text-align: center;
                background: white;
                padding: 20px 40px;
                border-radius: 8px;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                width: 100%;
                max-width: 500px;
            }
            form {
                margin-top: 20px;
            }
            input[type="file"] {
                padding: 10px;
                margin: 10px 0;
                width: 100%;
            }
            button {
                background-color: #007bff;
                color: white;
                border: none;
                padding: 10px 20px;
                cursor: pointer;
                border-radius: 4px;
                font-size: 16px;
            }
            button:hover {
                background-color: #0056b3;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>PDF to Excel Converter</h1>
            <form action="/upload" method="post" enctype="multipart/form-data">
                <input type="file" name="file" accept="application/pdf" required>
                <button type="submit">Convert to Excel</button>
            </form>
        </div>
    </body>
    </html>
    '''


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    # Save uploaded file
    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(file_path)

    try:

        # Convert PDF + images to Excel
        output_filename = file.filename.replace('.pdf', '.xlsx')
        extracted_data, excel_file_path = process_pdf_to_excel_with_images(
            pdf_path=file_path,
            output_filename=output_filename,
            fixer_api_url=FIXER_API_URL,
            fixer_api_key=FIXER_API_KEY,
            output_folder=OUTPUT_FOLDER,
            image_folder="./images"
        )

        return render_template_string(generate_response_html(extracted_data, excel_file_path))

    except Exception as e:
        return jsonify({"error": str(e)}), 500


IMAGE_FOLDER = os.path.join(OUTPUT_FOLDER, 'images')
os.makedirs(IMAGE_FOLDER, exist_ok=True)


def process_pdf_to_excel_with_images(pdf_path, output_filename, fixer_api_url, fixer_api_key, output_folder,
                                     image_folder):
    import os
    import re
    import io
    import hashlib
    import requests
    import pandas as pd
    from PIL import Image
    from PyPDF2 import PdfReader
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Font
    import fitz

    os.makedirs(output_folder, exist_ok=True)
    os.makedirs(image_folder, exist_ok=True)

    extracted_data = {"Libelle": None, "Department": None, "Object": None, "Table": []}
    compte_comptable_mapping = {
        "train": 625100, "plane": 625100, "parking": 625100, "taxi": 625100,
        "fuel": 625110, "peage": 625130, "entretien vehicule": 625140,
        "hotel": 625200, "repas restaurant": 625300, "reception": 625700,
        "affranchissement": 626000, "telephonie": 626100, "achats divers": 606300,
        "food": 625300, "other": 606300
    }

    try:
        response = requests.get(fixer_api_url, params={"access_key": fixer_api_key})
        response.raise_for_status()
        rates = response.json().get("rates", {})
    except Exception as e:
        print(f"Failed to fetch conversion rates: {e}")
        rates = {}

    try:
        reader = PdfReader(pdf_path)
    except Exception as e:
        print(f"Error reading PDF: {e}")
        return None, None

    for page in reader.pages:
        text = page.extract_text()
        if not text:
            continue

        name_match = re.search(r"NAME\s*:?[\s]*([A-Za-zÀ-ÿ\s]+?)\s+DEPARTMENT\s*:?[\s]*([A-Za-zÀ-ÿ\s]+)", text,
                               re.IGNORECASE)
        if name_match:
            extracted_data["Libelle"] = name_match.group(1).strip()
            dept_raw = name_match.group(2).strip()
            dept_clean = re.split(r"\s+OBJECT", dept_raw, maxsplit=1)[0].strip()
            extracted_data["Department"] = dept_clean

        lines = text.split('\n')
        merged_lines = []
        buffer = ""

        for line in lines:
            buffer += " " + line.strip()
            if re.search(r"\d{1,2}\s+\w+\s+\d{4}.*\d+.*(EUR|USD|MAD|TND).*Card", buffer, re.IGNORECASE):
                merged_lines.append(buffer.strip())
                buffer = ""
        if buffer:
            merged_lines.append(buffer.strip())

        expense_keywords = ["Food", "Taxi", "Hotel", "Train", "Plane", "Parking", "Fuel", "Peage", "Entretienvehicule", "Reception", "Affranchissement", "Telephonie", "Achatsdivers", "Other"]

        for merged_line in merged_lines:
            merged_line = re.sub(r'\s+', '', merged_line).strip()
            split_pattern = r'(?=(' + '|'.join(expense_keywords) + r'))'
            entries = re.split(split_pattern, merged_line)
            entry_blocks = []
            entry_labels = []
            for i in range(1, len(entries), 2):
                entry_blocks.append(entries[i + 1].strip())
                entry_labels.append(entries[i].strip())

            for idx, entry in enumerate(entry_blocks):
                match = re.search(
                    r"(?i)(\d{1,2})([A-Za-z]+)(\d{4})([\d.,]+)(?:B\.?commission([\d.,]+))?([A-Z]{3})(Card|offCard)([A-Z]+)",
                    entry
                )

                if match:
                    try:
                        day = match.group(1)
                        month = match.group(2)
                        year = match.group(3)
                        date = f"{day} {month} {year}"
                        frais = float(match.group(4).replace(',', '.'))
                        commission_str = match.group(5)
                        commission = float(commission_str.replace(',', '.')) if commission_str else 0.0
                        devis = match.group(6).upper()
                        card = match.group(7).strip().capitalize()
                        expense_topic = match.group(8).strip().upper()

                        for stop_word in ["AVOCARBON", "FRANCE", "TOTAL", "DATE", "PHONE"]:
                            if stop_word in expense_topic:
                                expense_topic = expense_topic.split(stop_word)[0]
                        expense_topic = expense_topic.upper().strip()

                        labelle = entry_labels[idx]
                        converted_value = frais / rates.get(devis, 1.0) if rates.get(devis) else frais
                        compte_comptable = compte_comptable_mapping.get(labelle.lower(), "Non défini")

                        row = [compte_comptable, labelle, date, frais, commission, devis, round(converted_value, 2), card, expense_topic]
                        if len(row) == 9:
                            extracted_data["Table"].append(row)
                    except Exception as e:
                        print(f"[WARN] Failed to parse entry: {entry} – {e}")

    valid_rows = [row for row in extracted_data["Table"] if isinstance(row, list) and len(row) == 9]

    df = pd.DataFrame(valid_rows, columns=[
        "Compte Comptable", "Libelle", "Date", "Montant en Devise", "B.commission",
        "Devise", "EUR", "Card", "Expense Topic"
    ]) if valid_rows else pd.DataFrame(columns=[
        "Compte Comptable", "Libelle", "Date", "Montant en Devise", "B.commission",
        "Devise", "EUR", "Card", "Expense Topic"
    ])

    output_path = os.path.join(output_folder, output_filename)

    doc = fitz.open(pdf_path)
    images = []
    extracted_hashes = set()

    for page_num in range(len(doc)):
        page = doc[page_num]
        image_blocks = []
        for img in page.get_images(full=True):
            xref = img[0]
            try:
                base_image = doc.extract_image(xref)
                image_bytes = base_image["image"]
                image_ext = base_image["ext"]
                rects = page.get_image_rects(xref)
                if not rects:
                    continue
                top_y = rects[0].y0
                img_preview = Image.open(io.BytesIO(image_bytes)).convert("L").resize((100, 100))
                img_hash = hashlib.md5(img_preview.tobytes()).hexdigest()
                if img_hash in extracted_hashes:
                    continue
                extracted_hashes.add(img_hash)
                image_blocks.append({"y": top_y, "bytes": image_bytes, "ext": image_ext})
            except Exception as e:
                print(f"Error extracting image from xref {xref}: {e}")
                continue
        image_blocks.sort(key=lambda x: x["y"])
        for img_data in image_blocks:
            filename = f"image_{len(images) + 1}.{img_data['ext']}"
            output_image_path = os.path.join(image_folder, filename)
            with open(output_image_path, "wb") as f:
                f.write(img_data["bytes"])
            images.append(output_image_path)

    image_filenames = [os.path.basename(img) for img in images]
    while len(image_filenames) < len(df):
        image_filenames.append("")
    df["Image Filename"] = image_filenames

    if not df.empty:
        total_row = {col: "" for col in df.columns}
        total_row["Libelle"] = "TOTAL"
        total_row["Montant en Devise"] = df["Montant en Devise"].sum()
        total_row["B.commission"] = df["B.commission"].sum()
        total_row["EUR"] = df["EUR"].sum()
        df.loc[len(df.index)] = total_row

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        meta_df = pd.DataFrame({
            "Field": ["Libelle", "Department"],
            "Value": [extracted_data["Libelle"], extracted_data["Department"]]
        })
        meta_df.to_excel(writer, sheet_name="Summary", index=False, startrow=0)
        df.to_excel(writer, sheet_name="Summary", index=False, startrow=5)

    try:
        workbook = load_workbook(output_path)
        for idx, row in enumerate(valid_rows):
            image_filename = os.path.basename(images[idx]) if idx < len(images) else f"Row_{idx + 1}"
            base_sheet_name = os.path.splitext(image_filename)[0][:31]
            sheet_name = base_sheet_name
            count = 1
            while sheet_name in workbook.sheetnames:
                sheet_name = f"{base_sheet_name[:28]}_{count}"
                count += 1
            sheet = workbook.create_sheet(title=sheet_name)
            headers = [
                "Compte Comptable", "Libelle", "Date", "Montant en Devise", "B.commission",
                "Devise", "EUR", "Card", "Expense Topic", "Image Filename"
            ]
            row_with_filename = row + [image_filename if idx < len(images) else ""]
            sheet.append(headers)
            sheet.append(row_with_filename)
            if idx < len(images):
                try:
                    sheet["A4"] = "Attached Receipt"
                    sheet["A4"].font = Font(bold=True)
                    img = ExcelImage(images[idx])
                    img.width, img.height = 300, 300
                    sheet.add_image(img, "A5")
                    sheet.row_dimensions[5].height = 200
                except Exception as e:
                    print(f"Failed to insert image for row {idx + 1}: {e}")
        workbook.save(output_path)
    except Exception as e:
        print(f"Failed to write image sheets: {e}")

    return extracted_data, output_path





def generate_response_html(extracted_data, excel_file_path):
    import os

    # Escape HTML special characters to prevent rendering issues
    def escape_html(text):
        return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    # Build table rows
    table_rows = "".join(
        f"<tr><td>{escape_html(row[0])}</td><td>{escape_html(row[1])}</td><td>{escape_html(row[2])}</td>"
        f"<td>{escape_html(row[3])}</td><td>{escape_html(row[4])}</td><td>{escape_html(row[5])}</td>"
        f"<td>{escape_html(row[6])}</td><td>{escape_html(row[7])}</td><td>{escape_html(row[8])}</td></tr>"
        for row in extracted_data.get("Table", [])
    )

    # Start building the HTML response
    return f'''
    <!doctype html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Extracted Data</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                background-color: #f4f4f9;
                margin: 0;
                padding: 20px;
            }}
            .container {{
                max-width: 900px;
                margin: auto;
                background: white;
                padding: 30px;
                border-radius: 10px;
                box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
            }}
            h1, h2 {{
                color: #333;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
            }}
            table, th, td {{
                border: 1px solid #ddd;
                text-align: left;
                padding: 10px;
            }}
            th {{
                background-color: #007bff;
                color: white;
                font-weight: bold;
            }}
            tr:nth-child(even) {{
                background-color: #f9f9f9;
            }}
            a.download-link {{
                display: inline-block;
                margin-top: 30px;
                text-align: center;
                color: white;
                background-color: #007bff;
                padding: 12px 20px;
                text-decoration: none;
                border-radius: 5px;
                font-size: 16px;
            }}
            a.download-link:hover {{
                background-color: #0056b3;
            }}
        </style>
    </head>
<div class="container">
    <h1>Extracted Data</h1>
    <p><strong>Name:</strong> {escape_html(extracted_data.get("Libelle", "N/A"))}</p>
    <p><strong>Department:</strong> {escape_html(extracted_data.get("Department", "N/A"))}</p>

    <h2>Expense Table</h2>
    <table>
        <thead>
            <tr>
                <th>Compte Comptable</th>
                <th>Libelle</th>
                <th>Date</th>
                <th>Montant en Devise</th>
                <th>B.commission</th>
                <th>Devise</th>
                <th>EUR</th>
                <th>Card</th>
                <th>Expense Topic</th>
            </tr>
        </thead>
        <tbody>
            {table_rows or "<tr><td colspan='9'>No data available.</td></tr>"}
        </tbody>
    </table>

    <a class="download-link" href="/download?file={os.path.basename(excel_file_path)}">⬇ Download Excel File</a>
</div>

    </html>
    '''


@app.route('/download')
def download_file():
    file_name = request.args.get('file')
    file_path = os.path.join(OUTPUT_FOLDER, file_name)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "File not found", 404


if __name__ == '__main__':
    app.run(debug=True)
